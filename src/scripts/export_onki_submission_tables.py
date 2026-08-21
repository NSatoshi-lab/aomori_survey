#!/usr/bin/env python3
"""Export manuscript Table 1 as a submission-ready English workbook."""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, Side


ITEM_LABELS = {
    "年齢": "Age",
    "住宅種別": "Housing type",
    "築年帯": "Building age",
    "所有形態": "Housing tenure",
    "入浴頻度": "Frequency of bathing at home in winter",
    "浴室窓": "Bathroom window",
    "浴室暖房乾燥機": "Bathroom heating dryer",
    "セントラル暖房": "Central heating",
    "その他暖房設備": (
        "Other equipment used to heat the dressing room or bathroom"
    ),
    "浴室寒さ体感": "Perceived bathroom coldness",
    "脱衣所寒さ体感": "Perceived dressing-room coldness",
}

CATEGORY_LABELS = {
    "18-49歳": "18–49 years",
    "50-59歳": "50–59 years",
    "60-69歳": "60–69 years",
    "70歳以上": "≥70 years",
    "一戸建て": "Detached house",
    "集合住宅": "Apartment/condominium",
    "30年未満": "<30 years",
    "30年以上": "≥30 years",
    "不明": "Unknown",
    "持家": "Owner-occupied",
    "賃貸": "Rented",
    "無回答": "Missing",
    "毎日": "Daily",
    "週4-6回": "4–6 times/week",
    "週1-3回": "1–3 times/week",
    "月1-3回": "1–3 times/month",
    "ほとんど入浴しない": "Almost never",
    "二重サッシ・複層ガラス": "Double window/double glazing",
    "単板ガラス": "Single glazing",
    "窓なし": "No window",
    "不明・無回答": "Unknown/missing",
    "設置して使用": "Installed and used for heating",
    "設置しているが未使用": "Installed but not used for heating",
    "未設置": "Not installed",
    "24時間使用": "Used 24 h/day (continuous)",
    "時間限定使用": "Used for limited hours",
    "不使用": "Not used",
    "ストーブ": "Space heater (stove)",
    "エアコン": "Air conditioner",
    "セントラル暖房以外の床暖房": (
        "Floor heating not part of the central heating system"
    ),
    "その他": "Other",
    "使用なし": "None",
    "5-7": "Coldness score 5–7",
}

EQUIPMENT_LABELS = {
    "浴室暖房乾燥機使用": "Bathroom heating dryer use",
    "浴室暖房乾燥機設置": "Bathroom heating dryer installation",
    "セントラル暖房使用": "Central heating use",
}

OUTCOME_LABELS = {
    "浴室寒さ5-7": "Bathroom",
    "脱衣所寒さ5-7": "Dressing room",
}

GROUP_LABELS = {
    "bath_heater_use": ("Not used", "Used"),
    "bath_heater_installed": ("Not installed", "Installed"),
    "central_heating": ("Not used", "Used"),
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("analysis_dir", type=Path)
    parser.add_argument("output_xlsx", type=Path)
    return parser.parse_args()


def format_ci(low: float, high: float) -> str:
    return f"[{low:.1f}, {high:.1f}]"


def build_table1(table1: pd.DataFrame) -> pd.DataFrame:
    def format_value(row: pd.Series) -> str:
        if row["item"] == "セントラル暖房" and row["category"] == "無回答":
            return str(int(row["event_n"]))
        return (
            f"{int(row['event_n'])} "
            f"({int(row['event_n']) / int(row['denominator']) * 100.0:.1f})"
        )

    characteristics = table1["item"].map(ITEM_LABELS)
    characteristics = characteristics.where(
        table1["item"].ne(table1["item"].shift()),
        "",
    )
    output = pd.DataFrame(
        {
            "Characteristic": characteristics,
            "Category": table1["category"].map(CATEGORY_LABELS),
            "n (%)": table1.apply(format_value, axis=1),
        }
    )
    if output[["Characteristic", "Category"]].isna().any().any():
        raise ValueError("Unmapped Table 1 label")
    return output


def build_table2(table2: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for row in table2.itertuples(index=False):
        comparison_label, equipment_label = GROUP_LABELS[row.equipment_key]
        rows.append(
            {
                "Equipment classification": EQUIPMENT_LABELS[row.equipment_label],
                "Area": OUTCOME_LABELS[row.outcome_label],
                "Comparison group n/N (%; 95% CI)": (
                    f"{comparison_label} "
                    f"{int(row.without_event_n)}/{int(row.without_denominator)} "
                    f"({row.without_pct:.1f}; "
                    f"{format_ci(row.without_ci95_lo_pct, row.without_ci95_hi_pct)})"
                ),
                "Equipment group n/N (%; 95% CI)": (
                    f"{equipment_label} "
                    f"{int(row.with_event_n)}/{int(row.with_denominator)} "
                    f"({row.with_pct:.1f}; "
                    f"{format_ci(row.with_ci95_lo_pct, row.with_ci95_hi_pct)})"
                ),
            }
        )
    return pd.DataFrame(rows)


def style_workbook(output_xlsx: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(output_xlsx)
    for sheet in workbook.worksheets:
        data_start_row = 4
        data_end_row = sheet.max_row - 2
        footnote_row = sheet.max_row
        thin_gray = Side(style="thin", color="B7B7B7")
        medium_black = Side(style="medium", color="000000")

        sheet.freeze_panes = "A4"
        sheet.merge_cells("A1:C1")
        sheet["A1"] = (
            "Table 1. Characteristics and use of heating equipment in the "
            "primary analysis sample (n = 147)"
        )
        sheet["A1"].font = Font(name="Times New Roman", size=11, bold=True)
        sheet["A1"].alignment = Alignment(
            wrap_text=True,
            vertical="center",
        )
        sheet.row_dimensions[1].height = 24

        for cell in sheet[3]:
            cell.font = Font(name="Times New Roman", size=10, bold=True)
            cell.alignment = Alignment(
                horizontal="right" if cell.column == 3 else "left",
                wrap_text=True,
                vertical="center",
            )
            cell.border = Border(
                top=medium_black,
                bottom=medium_black,
            )
        sheet.row_dimensions[3].height = 22

        for row_index in range(data_start_row, data_end_row + 1):
            group_start = bool(sheet.cell(row=row_index, column=1).value)
            for column_index in range(1, 4):
                cell = sheet.cell(row=row_index, column=column_index)
                cell.font = Font(
                    name="Times New Roman",
                    size=10,
                    bold=group_start and column_index == 1,
                )
                cell.alignment = Alignment(
                    horizontal="right" if column_index == 3 else "left",
                    indent=1 if column_index == 2 else 0,
                    wrap_text=True,
                    vertical="top",
                )
                cell.border = Border(
                    top=thin_gray if group_start else None,
                    bottom=medium_black if row_index == data_end_row else None,
                )
            characteristic = sheet.cell(row=row_index, column=1).value
            if (
                characteristic
                == "Other equipment used to heat the dressing room or bathroom"
            ):
                row_height = 38
            elif characteristic == "Frequency of bathing at home in winter":
                row_height = 42
            elif characteristic in {
                "Perceived bathroom coldness",
                "Perceived dressing-room coldness",
            }:
                row_height = 30
            else:
                row_height = 18
            sheet.row_dimensions[row_index].height = row_height

        sheet.cell(row=footnote_row, column=1).font = Font(
            name="Times New Roman",
            size=9,
        )
        sheet.cell(row=footnote_row, column=1).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )
        sheet.row_dimensions[footnote_row].height = 82

        sheet.column_dimensions["A"].width = 26
        sheet.column_dimensions["B"].width = 40
        sheet.column_dimensions["C"].width = 14
        sheet.sheet_view.showGridLines = False
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 1
        sheet.page_margins.left = 0.5
        sheet.page_margins.right = 0.5
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        sheet.print_options.horizontalCentered = True
        sheet.print_area = f"A1:C{sheet.max_row}"
        sheet.print_title_rows = "3:3"
    workbook.save(output_xlsx)


def main() -> None:
    args = parse_args()
    analysis_dir = args.analysis_dir.resolve()
    output_xlsx = args.output_xlsx.resolve()
    table1 = pd.read_csv(analysis_dir / "table1_characteristics.csv")
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        exported = build_table1(table1)
        exported.to_excel(
            writer,
            sheet_name="Table 1",
            index=False,
            startrow=2,
        )
        worksheet = writer.sheets["Table 1"]
        footnote_row = len(exported) + 5
        worksheet.cell(
            row=footnote_row,
            column=1,
            value=(
                "Values are presented as n (%). For central-heating use, 145 "
                "of the 154 returned questionnaires had no missing data; the 9 missing "
                "responses are shown in a separate row without a percentage. "
                "Percentages for all other characteristics were calculated using "
                "the 147 respondents in the primary analysis as the denominator. "
                "The use of other equipment to heat the dressing room or bathroom, "
                "excluding the bathroom heating dryer and central heating, was "
                "assessed using a multiple-response item; therefore, the percentages "
                "do not sum to 100%. The 8 missing responses for this item are "
                "shown in a separate row, with the percentage calculated using "
                "the 147 respondents as the denominator. For the bathroom and "
                "dressing room, perceived thermal sensation was rated on a "
                "7-point scale, from 1 (very warm) to 7 (very cold): the "
                "ratings were treated as the respective coldness scores; only "
                "the coldness score 5–7 group is shown for each room."
            ),
        )
        worksheet.merge_cells(
            start_row=footnote_row,
            start_column=1,
            end_row=footnote_row,
            end_column=3,
        )
    style_workbook(output_xlsx)
    print(output_xlsx)


if __name__ == "__main__":
    main()
